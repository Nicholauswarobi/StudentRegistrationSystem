import os
from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"

window = Tk()
window.title("Student Registration System")
window.geometry("1250x700+210+100")
window.config(bg=background)
# window.resizable(False, False)
icon = PhotoImage(file="icons/upload3.png")
window.iconphoto(False, icon)

file = pathlib.Path("Student_data.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Registration No"
    sheet["B1"] = "Name"
    sheet["C1"] = "Course"
    sheet["D1"] = "Gender"
    sheet["E1"] = "DOB"
    sheet["F1"] = "Date Of Registration"
    sheet["G1"] = "Religion"
    sheet["H1"] = "Skill"
    sheet["I1"] = "Father Name"
    sheet["J1"] = "Mother Name"
    sheet["K1"] = "Father's Occupation"
    sheet["L1"] = "Mother's Occupation"

    file.save("Student_data.xlsx")


##############################Exit window function###################################
def Exit():
    window.destroy()


###############################showimage############################################
def showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetypes=(("JPG File", "*.jpg"),
                                                                                ("PNG File", "*.png"),
                                                                                ("All files", "*.txt")))
    img = (Image.open(filename))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    img_label.config(image=photo2)
    img_label.image = photo2


##############################registration no########################################
def registration_no():
    global file
    global sheet
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("22100533590001")


####################################Clear####################################
def clear():
    global img
    Name.set("")
    DOB.set("")
    Religion.set("")
    Skill.set("")
    fname.set("")
    Mname.set("")
    FO.set("")
    MO.set("")
    Course.set("Select Course")

    registration_no()

    save_button.config(state="normal")

    img1 = PhotoImage(file="icons/upload3.png")
    img_label.config(image=img1)
    img_label.image = img1

    img = ""


##########################################Save########################################
def save():
    global G1
    global R1
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Course.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Please select Gender.")

    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = fname.get()
    mothername = Mname.get()
    F1 = FO.get()
    M1 = MO.get()

    if N1 == "" or C1 == "Select Course" or Rel == "" or D2 == "" or S1 == "" or fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("Error", "Few data is missing!")
    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=R1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=C1)
        sheet.cell(column=4, row=sheet.max_row, value=G1)
        sheet.cell(column=5, row=sheet.max_row, value=D2)
        sheet.cell(column=6, row=sheet.max_row, value=D1)
        sheet.cell(column=7, row=sheet.max_row, value=Rel)
        sheet.cell(column=8, row=sheet.max_row, value=S1)
        sheet.cell(column=9, row=sheet.max_row, value=fathername)
        sheet.cell(column=10, row=sheet.max_row, value=mothername)
        sheet.cell(column=11, row=sheet.max_row, value=F1)
        sheet.cell(column=12, row=sheet.max_row, value=M1)

        file.save(r'Student_data.xlsx')

        try:
            img.save("Students_Images/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("Info", "Profile picture is not available!!")
        messagebox.showinfo("Info", "Successfully data Entered!!")

        clear()
        registration_no()


######################################Search#########################################
def search():
    global reg_number
    global name
    text = Search.get()

    clear()

    save_button.config(state='disabled')
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]

    try:
        print(name)
    except:
        messagebox.showerror("Invalid", "Invalid registration number!!")

    x1 = sheet.cell(row=int(reg_number), column=1).value
    x2 = sheet.cell(row=int(reg_number), column=2).value
    x3 = sheet.cell(row=int(reg_number), column=3).value
    x4 = sheet.cell(row=int(reg_number), column=4).value
    x5 = sheet.cell(row=int(reg_number), column=5).value
    x6 = sheet.cell(row=int(reg_number), column=6).value
    x7 = sheet.cell(row=int(reg_number), column=7).value
    x8 = sheet.cell(row=int(reg_number), column=8).value
    x9 = sheet.cell(row=int(reg_number), column=9).value
    x10 = sheet.cell(row=int(reg_number), column=10).value
    x11 = sheet.cell(row=int(reg_number), column=11).value
    x12 = sheet.cell(row=int(reg_number), column=12).value

    Registration.set(x1)
    Name.set(x2)
    Course.set(x3)

    if x4 == "female":
        R2.select()
    else:
        R1.select()

    DOB.set(x5)
    Date.set(x6)
    Religion.set(x7)
    Skill.set(x8)
    fname.set(x9)
    Mname.set(x10)
    FO.set(x11)
    MO.set(x12)

    img = (Image.open("Students_Images/" + str(x1) + ".jpg"))
    resized_image = img.resize((190, 190))
    photo2 = ImageTk.PhotoImage(resized_image)
    img_label.config(image=photo2)
    img_label.image = photo2

######################################Update##########################################
def update():
    global G1
    global R1
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Course.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    Rel = Religion.get()
    S1 = Skill.get()
    fathername = fname.get()
    mothername = Mname.get()
    F1 = FO.get()
    M1 = MO.get()

    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            reg_no_position = str(name)[14:-1]
            reg_number = str(name)[15:-1]
            print(reg_number)

            # sheet.cell(column = 1, row = int(reg_number), value = R1)
            sheet.cell(column = 2, row = int(reg_number), value = N1)
            sheet.cell(column = 3, row = int(reg_number), value = C1)
            sheet.cell(column = 4, row = int(reg_number), value = G1)
            sheet.cell(column = 5, row = int(reg_number), value = D2)
            sheet.cell(column = 6, row = int(reg_number), value = D1)
            sheet.cell(column = 7, row = int(reg_number), value = Rel)
            sheet.cell(column = 8, row = int(reg_number), value = S1)
            sheet.cell(column = 9, row = int(reg_number), value = fathername)
            sheet.cell(column = 10, row = int(reg_number), value = mothername)
            sheet.cell(column = 11, row = int(reg_number), value = F1)
            sheet.cell(column = 12, row = int(reg_number), value = M1)

            file.save(r'Student_data.xlsx')

            try:
                img.save("Students_Images/" + str(R1) + ".jpg")
            except:
                pass
            messagebox.showinfo("Update", "Successfully data updated!!")

    clear()
    search()

######################################Gender function#################################
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"

    else:
        gender = "Female"


# Top frame
Label(window, text="Email: warobinicholaus@gmail.com", width=10, height=3,
      bg="#f0687c", anchor="e").pack(side=TOP, fill=X)
Label(window, text="STUDENT REGISTRATION", width=10, height=2,
      bg="#c36465", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)

# Search entry to update details
Search = StringVar()
Entry(window, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=68)
search_image = PhotoImage(file="icons/search1.png")
search_button = Button(window, text="Search", compound=LEFT, image=search_image, width=100, height=30,
                       bg="#68ddfa", font="arial 13 bold", command=search)
search_button.place(x=1060, y=70)

update_image = PhotoImage(file="icons/update2.png")
update_button = Button(window, text="Update", compound=LEFT, image=update_image,
                       width=100, font="arial 13 bold", bg="#68ddfa", command = update)
update_button.place(x=110, y=70)

# Registration and date
Label(window, text="Registration No:", font="arial 13 bold", bg=background, fg=framebg).place(x=30, y=150)
Label(window, text="Date:", font="arial 13 bold", bg=background, fg=framebg).place(x=500, y=150)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(window, width=15, textvariable=Registration)
reg_entry.place(x=170, y=150)

registration_no()

today = date.today()
date = today.strftime("%d/%m/%Y")
date_entry = Entry(window, width=15, textvariable=Date)
date_entry.place(x=550, y=150)
Date.set(date)

# Student details
student_frame = LabelFrame(window, text="Student's Details", font="arial 20 bold", bg=framebg, fg=framefg, height=250,
                           width=900, bd=2, relief=GROOVE)
student_frame.place(x=30, y=200)

Label(student_frame, text="Full Name:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=50)
Label(student_frame, text="Date of Birth:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=100)
Label(student_frame, text="Gender:", font="arial 13", fg=framefg, bg=framebg).place(x=30, y=150)

Label(student_frame, text="Course:", font="arial 13", fg=framefg, bg=framebg).place(x=500, y=50)
Label(student_frame, text="Religion:", font="arial 13", fg=framefg, bg=framebg).place(x=500, y=100)
Label(student_frame, text="Skills:", font="arial 13", fg=framefg, bg=framebg).place(x=500, y=150)

Name = StringVar()
name_entry = Entry(student_frame, textvariable=Name, width=30)
name_entry.place(x=150, y=50)

DOB = StringVar()
bod_entry = Entry(student_frame, textvariable=DOB, width=30)
bod_entry.place(x=150, y=100)

radio = IntVar()
R1 = Radiobutton(student_frame, text="Male", variable=radio, value=1, bg=framebg,
                 fg=framefg, command=selection)
R1.place(x=150, y=150)

R2 = Radiobutton(student_frame, text="Female", variable=radio, value=2, bg=framebg,
                 fg=framefg, command=selection)
R2.place(x=230, y=150)

Religion = StringVar()
religion_entry = Entry(student_frame, textvariable=Religion, width=30)
religion_entry.place(x=600, y=100)

Skill = StringVar()
skill_entry = Entry(student_frame, textvariable=Skill, width=30)
skill_entry.place(x=600, y=150)

Course = Combobox(student_frame, values=["Engineering In Data science", "Computer Science", "Computer Engineering",
                                         "ICT", "IT", "Telecommunication"], font="arial 10", width=23, state="r")
Course.place(x=600, y=50)
Course.set("Select Course")

# Parent details
parent_frame = LabelFrame(window, text="Parent's Details", font="arial 20 bold", bg=framebg, fg=framefg, height=210,
                          width=900, bd=2, relief=GROOVE)
parent_frame.place(x=30, y=480)

Label(parent_frame, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=50)
Label(parent_frame, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=120)

Label(parent_frame, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=50)
Label(parent_frame, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=120)

fname = StringVar()
FO = StringVar()
fname_entry = Entry(parent_frame, textvariable=fname, width=30)
fname_entry.place(x=150, y=50)

FO_entry = Entry(parent_frame, textvariable=FO, width=30)
FO_entry.place(x=150, y=120)

Mname = StringVar()
MO = StringVar()
Mname_entry = Entry(parent_frame, textvariable=Mname, width=30)
Mname_entry.place(x=620, y=50)

MO_entry = Entry(parent_frame, textvariable=MO, width=30)
MO_entry.place(x=620, y=120)

# Image
Image_frame = Frame(window, bd=5, bg="black", width=200, height=200, relief=GROOVE)
Image_frame.place(x=1000, y=150)

img = PhotoImage(file="icons/upload3.png")
img_label = Label(Image_frame, image=img, bg="black")
img_label.place(x=0, y=0)

# Button
upload_button = Button(window, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue",
                       command=showimage)
upload_button.place(x=1000, y=370)

save_button = Button(window, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=save)
save_button.place(x=1000, y=450)

reset_button = Button(window, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=clear)
reset_button.place(x=1000, y=530)

exit_button = Button(window, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit)
exit_button.place(x=1000, y=610)

window.mainloop()
